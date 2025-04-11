import random
import yaml
import openpyxl
import os
import pdb
from openpyxl import Workbook, load_workbook
from time import sleep
from typing import List, Dict, Tuple
from jogador.jogador import Jogador
from monstro.monstro import Monstro
from config.settings import Settings
from functions.cores import Cores

class Batalha:
    """
    Classe Batalha: responsável por gerenciar a batalha entre o jogador e o monstro.
    
    Atributos:
        jogador (Jogador): Instância do jogador.
        monstro (Monstro): Instância do monstro.
        continuar (bool): Indica se a batalha deve continuar.
        dados (dict): Dados do jogo, como dinheiro da banca e monstros disponíveis.
        settings (Settings): Configurações do jogo.
        plataformaComVantagem (bool): Indica se a plataforma tem vantagem ou não.
    
    Métodos:
        menu(): Exibe o menu inicial do jogo e inicia a batalha.
        iniciar_batalha(): Inicia a batalha entre o jogador e o monstro.
        sofrimentoDeDano(alvo: str): Aplica o dano ao alvo (jogador ou monstro).
        escolhaDeMonstro(plataformaComVantagem): Define o monstro a ser escolhido.
        definirVantagem(): Define se a plataforma tem vantagem ou não.
        abatimentoNoDinheiro(resultado: str): Aplica o abatimento no dinheiro da banca.
    """
    def __init__(self, jogador:Jogador, monstro:Monstro, dados:Dict[str, Dict[str, float]]):
        self.jogador = jogador
        self.monstro = monstro
        self.rodadas = 0
        self.quantidadeDeBoss = 0
        self.continuar: bool = True
        self.dados = dados
        self.settings = Settings()
        self.cores = Cores()
        self.plataformaComVantagem: bool = Batalha.definirVantagem(self)
        
    
    def menu(self):
        """
        Menu inicial do jogo, exibe as opções de jogo e inicia a batalha.
        """
        if self.settings.skip == False:
            # Tela inicial
            print('\n\n\n\n')
            print(self.cores.yellow + '{:¨^41}'.format('¨'))
            print('{:¨^41}'.format('| Bem-vindo |'))
            print('{:¨^41}'.format('| Ao |'))
            print('{:¨^41}'.format('| JOGO DO PODER |'))
            print('{:¨^41}'.format('¨'))
            print('{:-^41}'.format('-'))
            print('\n\n')
            print('{:^41}'.format('Deseja entrar na Dungeon do Boss?'))
            print(self.cores.green + '                 Sim' + self.cores.yellow + ' / ' + self.cores.red + 'Não              ' + self.cores.normal)
        
        # Checando variável de teste
        if self.settings.teste == True:
            iniciar_jogo = 'SIM'
        else:
            iniciar_jogo = str(input('    --> ')).upper().strip()
            
        if iniciar_jogo == 'SIM' or iniciar_jogo == 'S':
            # Inserção de Dinheiro
            print(self.cores.yellow + '    MAX - R$ 100,00' + self.cores.normal)
            print(self.cores.green + '    Qual valor deseja apostar?')
            
            # Checando variável de teste
            if self.settings.teste == True:
                self.apostado = random.randint(10, 100)
            else:
                self.apostado = float(input('    R$' + self.cores.normal))
            print('\n')
            
    def iniciar_batalha(self):
        """
        inicia a batalha entre o jogador e o monstro.
        """
        while self.continuar == True:
            if self.settings.skip == False:
                sleep(self.settings.velocidadeDoJogo) # Velocidade do Jogo
                # Informação na tela
                print(self.cores.green + f'   {self.jogador.name}', end=' ')
                print(self.cores.yellow + f'Lv {self.jogador.level}' + self.cores.normal + ' X ' + self.cores.red + f'{self.monstro.name}', end=' ')
                print(self.cores.yellow + f'Lv {self.monstro.level}')
                print(f'   Power: {self.jogador.power:.2f}' + self.cores.normal, end='   X')
                print(self.cores.yellow + f'   Power: {self.monstro.power:.2f}' + self.cores.normal)

            # Resultado do round
            if self.jogador.power > self.monstro.power:
                print(self.cores.green + '  -x-x-x-x- Usuário Win -x-x-x-x-' + self.cores.normal + '\n')
                Batalha.sofrimentoDeDano(self, 'monstro')
                
            elif self.jogador.power < self.monstro.power:
                print(self.cores.red + '  -x-x-x-x- Monstro Win -x-x-x-x-' + self.cores.normal + '\n')
                Batalha.sofrimentoDeDano(self, 'jogador')
                
            else:
                print(self.cores.ciano + '  -x-x-x-x- Empate -x-x-x-x-' + self.cores.normal + '\n')
                Batalha.sofrimentoDeDano(self, 'ambos')

            self.rodadas += 1
    
    def sofrimentoDeDano(self, alvo: str):
        """
        Aplica o dano ao alvo (jogador ou monstro) e verifica o resultado da batalha.
        """
        if alvo == 'monstro':
            # Dano no monstro
            self.monstro.power -= (self.jogador.power / self.settings.danoAoMonstro)
            self.jogador.power += self.settings.aumentoDePoder
            self.jogador.level += 1
            
            if self.monstro.power <= 0:
                Batalha.escolhaDeMonstro(self, self.plataformaComVantagem)
                
        elif alvo == 'jogador':
            # Dano no jogador
            self.jogador.power -= (self.monstro.power / self.settings.danoAoJogador)
            # Refletindo o dano no monstro
            self.monstro.power -= (self.monstro.power / self.settings.danoRefletidoAoMonstro)
            # Aumento de Nível do Monstro (Mudará apenas o nível)
            self.monstro.level += 1
            
            if self.jogador.power <= 0:
                print(self.cores.yellow + '   -x-x-x-x--x-x-x-x--x-x-x-x--x-' + self.cores.normal)
                print(self.cores.yellow + '   -x-x-x-x-' + self.cores.red + 'Você Morreu!' + self.cores.yellow + '-x-x-x-x-' + self.cores.normal)
                print(self.cores.yellow + '   -x-x-x-x--x-x-x-x--x-x-x-x--x-' + self.cores.normal)
                print(self.cores.yellow + '\n   Valor apostado: ' + self.cores.red + '-R$' + str(self.apostado))
                print(self.cores.normal)
                Batalha.abatimentoNoDinheiro(self, 'derrota')
        else:
            # Empate
            self.jogador.power -= (self.monstro.power / self.settings.danoAoJogador)
            self.monstro.power -= (self.jogador.power / self.settings.danoAoMonstro)
            if self.jogador.power <= 0:
                print(f'O {self.jogador.name} foi derrotado!')
            elif self.monstro.power <= 0:
                print(f'O {self.monstro.name} foi derrotado!')
            elif self.jogador.power <= 0 and self.monstro.power <= 0:
                print('Ambos os lados foram derrotados!')
                Batalha.abatimentoNoDinheiro(self, 'empate')
        
        if self.jogador.level >= 100:
            print('\n\n' + self.cores.green + 'Parabéns!')
            print('   Você venceu a Dungeon!' + self.cores.normal)
            print(self.cores.green + '   Valor Acumulado R${:.2f}'.format(self.apostado))
            print(self.cores.normal)
            Batalha.abatimentoNoDinheiro(self, 'vitoria')
    
    def escolhaDeMonstro(self, plataformaComVantagem):
        """
        Definição do monstro a ser escolhido, levando em consideração a da plataforma.
        """
        monstros: List[str] = self.dados['monstros']
        
        # Geração aleatória do Boss Infernal
        if self.monstro.name != 'Boss Infernal':
            aparicao = self.monstro.bossInfernal()
            if aparicao == True:
                self.monstro.name = self.monstro.name
                self.monstro.power = self.monstro.power
                self.monstro.level = self.monstro.level
                
                sleep(self.settings.velocidadeDoJogoBoss) # Velocidade da aparição do Boss
                print(self.cores.red + f'   VOCÊ ENCONTROU UM BOSS INFERNAL NIVEL {self.monstro.level}' + self.cores.normal)
                sleep(self.settings.velocidadeDoJogo)
                
                self.quantidadeDeBoss += 1
                
                return
        
        if plataformaComVantagem == True:
            monstroEscolhido: str = random.choice(list(monstros.keys()))
            monstroPower: float = monstros[monstroEscolhido]
            self.monstro.name = monstroEscolhido
            self.monstro.power = monstroPower
            
            # Tenta garantir a vantagem com um monstro forte
            for _ in range(20):
                if monstroPower < self.jogador.power:
                    monstroEscolhido = random.choice(list(monstros.keys()))
                    monstroPower = monstros[monstroEscolhido]
                    self.monstro.name = monstroEscolhido
                    self.monstro.power = monstroPower
        
        else:
            monstroEscolhido = random.choice(list(monstros.keys()))
            monstroPower = monstros[monstroEscolhido]
            self.monstro.name = monstroEscolhido
            self.monstro.power = monstroPower

    def definirVantagem(self):
        """
        Define se a plataforma tem vantagem ou não, baseado no valor da banca e da variável de vantagemMaster.
        """
        plataformaComVantagem: bool = None # Variável para definir a vantagem do jogo
        
        if self.dados['banca']['dinheiro'] < self.settings.valorMinimo and self.settings.vantagemMaster == True:
            if random.Random().randint(1, 6) >= 3:
                plataformaComVantagem = True               
            else:
                plataformaComVantagem = False
        else:
            plataformaComVantagem = False
            
        return plataformaComVantagem

    def abatimentoNoDinheiro(self, resultado: str):
        """
        Aplica o abatimento no dinheiro da banca baseado no resultado da batalha.
        """
        if resultado == 'vitoria':
            self.dados['banca']['dinheiro'] -= self.apostado
        elif resultado == 'derrota':
            self.dados['banca']['dinheiro'] += self.apostado
        else:
            self.dados['banca']['dinheiro'] = self.dados['banca']['dinheiro'] # Valor padrão
        
        # Coletar e Exportar dados
        self.exportarDados(resultado)
        self.continuar = False
        
        # Atualiza o arquivo gameData.yaml
        with open(self.settings.caminhoGameData, 'w', encoding='utf-8') as arquivo:
            yaml.dump(self.dados, arquivo, default_flow_style=False, allow_unicode=True)
            
    def exportarDados(self, resultado: str = None):
        try:
            # Caminho para a pasta e o arquivo
            pasta = "data"
            caminho_arquivo = os.path.join(pasta, "relatorio_de_dados.xlsx")

            # Cria a pasta se não existir
            os.makedirs(pasta, exist_ok=True)

            # Verifica se o arquivo já existe
            if os.path.exists(caminho_arquivo):
                wb = load_workbook(caminho_arquivo)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Relatório de Dados"
                ws.append(["Nome", "Nível Máximo", "Rodadas", "Valor Apostado", "Qnt Bosses", "Vantagem", "Resultado"])

            # Adiciona os dados
            ws.append([self.jogador.name, self.jogador.level, self.rodadas, self.apostado, self.quantidadeDeBoss, self.plataformaComVantagem, resultado.upper()])
            wb.save(caminho_arquivo)
        except Exception as e:
            print(f"Erro ao coletar dados: {e}")